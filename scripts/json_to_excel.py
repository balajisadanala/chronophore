#!/usr/bin/python3

import argparse
import json
import logging
import openpyxl
from collections import OrderedDict


def data_to_excel(data, output_file):
    """Saves data from one of Chronophore's json files
    to an Excel spreadsheet.

    Parameters:
        - data: A dictionary (preferably an OrderedDict) from
                a json file created by Chronophore.
        - title: The title of the output file.
    """
    try:
        headers = {header for key in data for header in data[key].keys()}
        logging.debug("Headers: {}".format(headers))
    except (TypeError, AttributeError) as e:
        logging.error(
            "Invalid data. Must be a valid dictionary: {} ({})".format(data, e)
        )
    else:
        # assign columns to headers
        header_columns = {
            header: column for key in data
            for column, header in enumerate(headers, start=2)
        }

        # create sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title, _ = output_file.split('.')

        # make headers
        sheet.cell(row=1, column=1).value = "Key"
        for (header, col_num) in header_columns.items():
            sheet.cell(row=1, column=col_num).value = header

        sheet.freeze_panes = 'A2'

        # fill in data
        for row_num, key in enumerate(data.keys(), start=2):
            sheet.cell(row=row_num, column=1).value = key
            for header, value in data[key].items():
                col_num = header_columns[header]
                sheet.cell(row=row_num, column=col_num).value = value

        wb.save(output_file)
        logging.info("Worksheet saved: {}".format(output_file))


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description="Convert Chronophore data from json to xlsx."
    )
    parser.add_argument('input', help="path of json file to use as input")
    parser.add_argument(
        '-o', '--output',
        help="path of excel file to generate (default: $(input_file_name).xls)"
    )
    args = parser.parse_args()

    json_file = args.input
    name, ext = json_file.split('.')

    excel_file = args.output if args.output else (name + '.xlsx')

    with open(json_file, 'r') as f:
        data = json.load(f, object_pairs_hook=OrderedDict)

    data_to_excel(data, excel_file)
